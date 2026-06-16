#!/usr/bin/env python3
from __future__ import annotations

import json
import random
import re
from pathlib import Path
from typing import Any, Iterable

import openpyxl


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "generated_common_name_cases"
OUTPUT_FILE = OUTPUT_DIR / "ordinary_name_privacy_cases.json"
NAME_FILE = BASE_DIR / "name" / "name.txt"

RANDOM_SEED = 20260526

COUNTRIES = [
    {
        "label": "沙特",
        "id_prefix": "saudi",
        "data_file": BASE_DIR / "data" / "隐私数据项-沙特.xlsx",
        "template_file": BASE_DIR / "template" / "ar.txt",
        "name_mode": "common_pair",
    },
    {
        "label": "泰国",
        "id_prefix": "thailand",
        "data_file": BASE_DIR / "data" / "隐私数据项-泰国.xlsx",
        "template_file": BASE_DIR / "template" / "th.txt",
        "name_mode": "given_surname",
    },
    {
        "label": "土耳其",
        "id_prefix": "turkey",
        "data_file": BASE_DIR / "data" / "隐私数据项-土耳其.xlsx",
        "template_file": BASE_DIR / "template" / "tu.txt",
        "name_mode": "given_surname",
    },
]

SECTION_HEADERS = {"中文", "英文", "土耳其", "沙特", "泰国"}
TERM_HEADER = "翻译"


def main() -> int:
    rng = random.Random(RANDOM_SEED)
    name_text = NAME_FILE.read_text(encoding="utf-8")

    all_records: list[dict[str, Any]] = []
    summary: dict[str, int] = {}

    for country in COUNTRIES:
        label = country["label"]
        templates = load_templates(country["template_file"])
        terms = load_terms(country["data_file"])
        names = load_name_pool(name_text, label, country["name_mode"])

        country_records = build_records(
            country_label=label,
            start_id=len(all_records) + 1,
            templates=templates,
            terms=terms,
            names=names,
            rng=rng,
        )
        all_records.extend(country_records)
        summary[label] = len(country_records)

    validate_records(all_records)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(
        json.dumps(all_records, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(json.dumps({"output": str(OUTPUT_FILE), "summary": summary, "total": len(all_records)}, ensure_ascii=False, indent=2))
    return 0


def load_templates(path: Path) -> list[str]:
    templates = [line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
    if not templates:
        raise ValueError(f"No templates found in {path}")
    for template in templates:
        if "【sth】" not in template or "【sb】" not in template:
            raise ValueError(f"Template missing placeholders in {path}: {template}")
    return templates


def load_terms(path: Path) -> list[tuple[str, str]]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    headers = [clean_text(sheet.cell(1, col).value) for col in range(1, sheet.max_column + 1)]
    try:
        item_col = headers.index("隐私数据项") + 1
        term_col = headers.index(TERM_HEADER) + 1
    except ValueError as exc:
        raise ValueError(f"No privacy item or translation column found in {path}") from exc

    terms: list[tuple[str, str]] = []
    for row in range(2, sheet.max_row + 1):
        item_name = clean_term(sheet.cell(row, item_col).value)
        term = clean_term(sheet.cell(row, term_col).value)
        if item_name and term:
            terms.append((item_name, term))

    if not terms:
        raise ValueError(f"No data terms found in {path}")
    return terms


def load_name_pool(name_text: str, country_label: str, mode: str) -> list[str]:
    section = extract_country_section(name_text, country_label)
    common_text = section.split("公众人物", 1)[0]

    if mode == "given_surname":
        given_text = extract_labeled_block(common_text, "名：", "姓：")
        surname_text = extract_labeled_block(common_text, "姓：", None)
        given_names = split_name_tokens(given_text)
        surnames = split_name_tokens(surname_text)
        if not given_names or not surnames:
            raise ValueError(f"Could not extract given names and surnames for {country_label}")
        return [f"{given} {surname}" for given in given_names for surname in surnames]

    if mode == "common_pair":
        common_names = split_name_tokens(common_text)
        if len(common_names) < 2:
            raise ValueError(f"Need at least two common names for {country_label}")
        return [f"{first} {last}" for first in common_names for last in common_names if first != last]

    raise ValueError(f"Unsupported name mode: {mode}")


def extract_country_section(text: str, country_label: str) -> str:
    lines = text.splitlines()
    start = None
    for idx, line in enumerate(lines):
        if line.strip() == country_label:
            start = idx + 1
            break
    if start is None:
        raise ValueError(f"Country section not found: {country_label}")

    end = len(lines)
    for idx in range(start, len(lines)):
        if lines[idx].strip() in SECTION_HEADERS:
            end = idx
            break
    return "\n".join(lines[start:end])


def extract_labeled_block(text: str, start_label: str, end_label: str | None) -> str:
    start = text.find(start_label)
    if start < 0:
        return ""
    start += len(start_label)
    if end_label is None:
        end = text.find("公众人物", start)
    else:
        end = text.find(end_label, start)
    if end < 0:
        end = len(text)
    return text[start:end]


def split_name_tokens(text: str) -> list[str]:
    text = re.sub(r"常用人名（\d+）", "", text)
    text = re.sub(r"^(名|姓)：", "", text, flags=re.MULTILINE)
    tokens = re.split(r"[、,\n]+", text)
    return dedupe(clean_name_token(token) for token in tokens)


def build_records(
    *,
    country_label: str,
    start_id: int,
    templates: list[str],
    terms: list[tuple[str, str]],
    names: list[str],
    rng: random.Random,
) -> list[dict[str, Any]]:
    records = []
    for offset, (item_name, term) in enumerate(terms):
        template = templates[offset % len(templates)]
        person = rng.choice(names)
        case = template.replace("【sth】", term).replace("【sb】", person)
        records.append(
            {
                "id": start_id + offset,
                "item": f"{country_label}_{item_name}_{person}",
                "case": case,
            }
        )
    return records


def validate_records(records: list[dict[str, Any]]) -> None:
    ids = [record["id"] for record in records]
    if len(ids) != len(set(ids)):
        raise ValueError("Duplicate record ids found")

    bad_records = [
        record["id"]
        for record in records
        if set(record) != {"id", "item", "case"}
        or not isinstance(record["id"], int)
        or not record["item"]
        or "【sth】" in record["case"]
        or "【sb】" in record["case"]
    ]
    if bad_records:
        raise ValueError(f"Invalid records: {bad_records[:10]}")


def clean_term(value: object) -> str:
    text = clean_text(value)
    text = text.strip("\"'“”‘’")
    text = re.sub(r"\s*[—–-]\s*$", "", text)
    return text.strip()


def clean_name_token(value: object) -> str:
    text = clean_text(value)
    text = re.sub(r"^[：:]+", "", text)
    return text.strip("\"'“”‘’ \u202f")


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\u00a0", " ").replace("\u202f", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def dedupe(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


if __name__ == "__main__":
    raise SystemExit(main())
